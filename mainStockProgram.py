import openpyxl
import string
from download_fin_company import download_finnacials, delete_company_reports


class Company:
    def __init__(self, ticker_name):
        self.tickerName = ticker_name

    def __str__(self):
        print(self.tickerName)

    # download reports and open excel report work_books to manipulate
    def initiate_params(self):
        self.raw_Data_cagar = {}
        self.growthRate = {}
        self.ticker = self.tickerName

        # download company financials reports from www.stockrow.com
        info_messg = download_finnacials(self.ticker)
        if info_messg == "wrong ticker name":
            self.tickerName = info_messg
        else:
            balance_sheet_path = "C:/Users/liela/PycharmProjects/valueInvesting/stock/financials/{0}/balanceSheet.xlsx".format(self.ticker)
            income_statment_path = "C:/Users/liela/PycharmProjects/valueInvesting/stock/financials/{0}/incomeStatement.xlsx".format(self.ticker)
            metrics_path = "C:/Users/liela/PycharmProjects/valueInvesting/stock/financials/{0}/metrics.xlsx".format(self.ticker)

            # create financials workBook
            self.w_bs = openpyxl.load_workbook(balance_sheet_path)
            self.w_is = openpyxl.load_workbook(income_statment_path)
            self.w_mt = openpyxl.load_workbook(metrics_path)
            # open financials workbooks
            self.sheet_bs = self.w_bs.active
            self.sheet_is = self.w_is.active
            self.sheet_mt = self.w_mt.active
            # if company not exist at least  years in the market - there will be not enough data to evaluate it
            if self.sheet_bs.max_column < 9 or self.sheet_is.max_column < 9 or self.sheet_mt.max_column < 9:
                messg = "no enough data"
                self.tickerName = messg

    def get_financials_param(self, sheet_type, value_name):
        values = []
        cagar = []
        for i in range(1, sheet_type.max_row + 1):
            cell = "A" + str(i)
            if sheet_type[cell].value == value_name: # if report parameter exist in row
                for col in range(1, 9):
                    char = string.ascii_uppercase[col]
                    cell2 = char + str(i)
                    values.append(float(sheet_type[cell2].value))
                    if col % 2 == 0:
                        cagar_value_by_precents = ((values[0] / values[col - 1]) ** (1 / (col - 1)) - 1) * 100
                        cagar.append(int(cagar_value_by_precents.real))
                self.raw_Data_cagar[value_name] = values
                self.growthRate[value_name] = cagar
                break

    def close_wb(self):
        self.w_bs.close()
        self.w_is.close()
        self.w_mt.close()
        delete_company_reports(self.ticker)
